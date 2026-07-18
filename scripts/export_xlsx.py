#!/usr/bin/env python3
"""Export the source-of-truth CSV to the downloadable spreadsheet.

Reads ``data/blocks.csv`` and writes ``Minecraft_Redstone_Block_Index.xlsx``
with a formatted "Block Index" sheet (frozen header, auto-filter) plus a
"Read Me" sheet documenting the columns. The CSV remains the source of truth;
this is a generated artifact.

Usage:
    python scripts/export_xlsx.py
"""
import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / "data" / "blocks.csv"
XLSX_PATH = ROOT / "Minecraft_Redstone_Block_Index.xlsx"

# (csv key, spreadsheet header, column width)
COLUMNS = [
    ("name", "Name", 46),
    ("category", "Category", 26),
    ("piston", "Piston Behavior", 14),
    ("breaks_water", "Breaks From Water", 12),
    ("stable", "Stable", 9),
    ("spawn_hostile", "Mob Spawnable (Hostile)", 12),
    ("spawn_friendly", "Mob Spawnable (Friendly)", 12),
    ("solid", "Solid", 9),
    ("conductive", "Conductive (Redstone)", 12),
    ("height", "Height (px)", 10),
    ("height_note", "Height Notes", 40),
    ("width", "Width/Length (px)", 12),
    ("width_note", "Width/Length Notes", 40),
    ("falls", "Falls (Gravity)", 10),
    ("flammable", "Flammable", 10),
    ("notes", "Notes", 70),
    ("id", "ID", 30),
]

READ_ME = [
    ("title", "Minecraft Redstone Block Index"),
    ("subtitle", "Java Edition \u2014 reference data for movability, water/piston interactions, "
                 "stability, mob spawning, redstone conductivity, collision hitbox size, and physics"),
    ("h", "What this is"),
    ("p", "A catalogue of Minecraft (Java Edition) blocks organized by BEHAVIOR FAMILY \u2014 blocks that "
          "share identical properties (e.g. all 16 wool colors, or every stair regardless of material) are "
          "combined into a single row so the sheet stays focused on mechanics rather than repeating "
          "near-identical rows."),
    ("p", "This spreadsheet is a GENERATED export. The source of truth is data/blocks.csv in the project "
          "folder \u2014 edit that CSV, then run `python scripts/build.py` (updates the web app) and "
          "`python scripts/export_xlsx.py` (regenerates this file)."),
    ("p", "Use the filter arrows on the header row (or the companion web app) to find blocks matching a "
          "combination of criteria, e.g. STABLE = TRUE and MOVABLE = FALSE."),
    ("h", "Columns"),
    ("p", "Piston Behavior \u2014 what happens when a piston tries to move this block. One of three values: "
          "MOVABLE (relocates intact, and can be dragged by slime/honey blocks), IMMOVABLE (the piston refuses "
          "to extend at all \u2014 nothing moves or breaks), or BREAKS (the piston still extends, but this block "
          "is destroyed instead of relocated)."),
    ("p", "Breaks From Water \u2014 does flowing water wash it away (pop it as a dropped item)?"),
    ("p", "Stable \u2014 can it exist without a specific supporting block below/above/beside it? "
          "(FALSE = it needs a specific attachment, like a torch needing a wall.)"),
    ("p", "Mob Spawnable (Hostile) \u2014 does this block's top surface meet the general rule for hostile mob "
          "spawning (solid + opaque + full top face), assuming light level and dimension otherwise allow it?"),
    ("p", "Mob Spawnable (Friendly) \u2014 same structural test, for passive/neutral mobs (natural overworld "
          "generation of animals is further restricted to Grass Block specifically; see its notes)."),
    ("p", "Solid \u2014 does it have a real collision hitbox a player can stand on / can't walk through?"),
    ("p", "Conductive (Redstone) \u2014 can the block be powered and pass a redstone signal through itself? "
          "Opaque full blocks (stone, wood, most ores) generally can; transparent blocks (glass, ice) and "
          "partial/non-full blocks (slabs, stairs, fences) generally can't. Note the named exceptions that go "
          "against the grain: Glowstone, Observer, Piston, TNT, Hopper and Honey Block (Java) are non-conductive "
          "despite being full cubes; Soul Sand and Mud are conductive despite not being full height."),
    ("p", "Height (px) \u2014 COLLISION HITBOX height out of 16px per block (1 block = 16px). This is the real "
          "collision box, not the visual model, wherever the two differ. Values above 16 (e.g. fences) extend "
          "above the block for collision purposes."),
    ("p", "Height Notes \u2014 clarifies stage-dependent or non-obvious height values."),
    ("p", "Width/Length (px) \u2014 COLLISION HITBOX horizontal footprint out of 16px. For blocks whose width and "
          "length differ (fences, doors, anvils, etc.) this is a representative value; see Width/Length Notes."),
    ("p", "Width/Length Notes \u2014 clarifies footprint values, especially anisotropic or stage-dependent ones."),
    ("p", "Falls (Gravity) \u2014 is it affected by gravity (becomes a Falling Block entity when unsupported)?"),
    ("p", "Flammable \u2014 can it catch fire and burn away? Simplified indicator covering the common, "
          "well-documented cases (wood, wool, leaves, vines, etc.); treat as a starting point, not exhaustive."),
    ("p", "Notes \u2014 mechanics, exceptions, and support requirements specific to that row."),
    ("p", "ID \u2014 stable slug used to join this row to the web app."),
    ("h", "Collision hitbox vs. visual model"),
    ("p", "Height and Width/Length are the block's COLLISION box, which can be smaller than what you see. "
          "Example: a Honey Block looks like a full cube but its collision box is 14\u00d714\u00d715px "
          "(width\u00d7length\u00d7height) \u2014 the 2px horizontal gap is why arrows squeeze between adjacent "
          "honey blocks. Soul Sand and Mud are 16\u00d714\u00d716 (top sits 2px low); slabs are 16px tall visually "
          "but 8px collision; and so on."),
    ("h", "The three PISTON BEHAVIOR states"),
    ("p", "Every block falls into exactly one state when a piston tries to move it (this replaces the old "
          "separate Movable / Breaks-From-Piston columns):"),
    ("p", "  1. MOVABLE \u2014 relocates intact (e.g. Stone). These are also the blocks slime/honey blocks can drag."),
    ("p", "  2. BREAKS \u2014 the piston still extends, but this block is destroyed instead of relocated "
          "(e.g. Torch, Cobweb, Dragon Egg, most attached/plant blocks)."),
    ("p", "  3. IMMOVABLE \u2014 the piston refuses to extend at all; nothing moves and nothing breaks "
          "(e.g. Obsidian, Chest, Furnace \u2014 mostly 'heavy'/technical blocks and anything storing an "
          "inventory)."),
    ("h", "Sourcing & methodology"),
    ("p", "Built from the Minecraft Wiki (minecraft.wiki) pages for Piston, Falling Block, Water, Mob spawning, "
          "Hitbox, and Conductivity, cross-checked against established Java Edition redstone mechanics, as of "
          "mid-2026 (game version ~1.21.x)."),
    ("p", "This is a reference tool, not a game data dump \u2014 a handful of rare/decorative blocks with unusual "
          "shapes (heads, dripstone, azalea, etc.) use approximate hitbox values, flagged in their Notes."),
    ("p", "Scope: placeable BLOCKS only. Entities that occupy block-like space (Item Frames, Paintings, Armor "
          "Stands, Boats, Minecarts) are excluded since they aren't blocks. Bedrock Edition differs on several of "
          "these mechanics (piston pushing of chests/hoppers, snow layer gravity, honey block conductivity, etc.) "
          "and is NOT covered here."),
    ("p", "Always double-check truly load-bearing technical builds in-game \u2014 Minecraft has version-to-version "
          "quirks and edge cases no reference sheet fully captures."),
]

HEADER_FILL = PatternFill("solid", fgColor="26211C")
HEADER_FONT = Font(color="ECDFC4", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="C73A2F")
SUB_FONT = Font(italic=True, size=11, color="555555")
SECTION_FONT = Font(bold=True, size=12, color="1F1A15")


def load_rows():
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def build_index_sheet(ws, rows):
    ws.title = "Block Index"
    headers = [h for _, h, _ in COLUMNS]
    ws.append(headers)
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    numeric_keys = {"height", "width"}
    for row in rows:
        record = []
        for key, _, _ in COLUMNS:
            val = row.get(key, "")
            if key in numeric_keys:
                try:
                    val = int(float(val))
                except (TypeError, ValueError):
                    val = 0
            record.append(val)
        ws.append(record)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 30


def build_readme_sheet(ws):
    ws.column_dimensions["B"].width = 120
    r = 2
    for kind, text in READ_ME:
        if kind == "h":
            r += 1  # blank line for breathing room before a section header
        cell = ws.cell(row=r, column=2, value=text)
        if kind == "title":
            cell.font = TITLE_FONT
        elif kind == "subtitle":
            cell.font = SUB_FONT
        elif kind == "h":
            cell.font = SECTION_FONT
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


def main():
    rows = load_rows()
    wb = openpyxl.Workbook()
    build_readme_sheet(wb.active)
    wb.active.title = "Read Me"
    build_index_sheet(wb.create_sheet(), rows)
    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH.name}: {len(rows)} blocks.")


if __name__ == "__main__":
    main()
